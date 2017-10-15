from bs4 import BeautifulSoup
from openpyxl import load_workbook
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from geopy.distance import vincenty
import requests, time, re, json, csv
import unicodedata

start_time = time.time()
date = time.strftime("%m/%d/%y")

def getWebData(link,headers,payload):
    return BeautifulSoup(requests.post(link,headers=headers,data=payload).text,'html.parser')

def remSpCh(s):
	if s is None:
		return ''
	s = unicodedata.normalize('NFKD', s).encode('ascii','ignore')
	return ''.join([i for i in s if ord(i) < 128])

def soupTextIfNotNone(obj):
	if obj is not None:
		return remSpCh(obj.text)
	return remSpCh(obj)

def getSheet(wb,sheetname):
	if sheetname in wb.get_sheet_names():
		return wb[sheetname]
	ws = wb.create_sheet()
	ws.title = sheetname
	return ws

def getEmptyCol(start_col, sheet):
	for i in range(start_col,999):
		val = sheet[get_column_letter(i)+'1'].value
		if val is None or val == time.strftime("%m/%d/%y"):
			return get_column_letter(i)

def getStartIndex(sheet):
	for i in range(2,sheet.max_row+2):
		val = sheet['B'+str(i)].value
		if val is None:
			return i
	return i

def get_zipcodes(sheet):
	zipcodes = []
	for i in range(1,sheet.max_row+1):
		val = sheet['A'+str(i)].value
		if val is None:
			break
		else:
			val = str(val)
			if len(val.split('-')[0].strip()) < 5:
				val = '0' + val
			zipcodes.append(val)
	return zipcodes

def getChanges(sheet,data,startIndex):
	listOfDoctors = []
	deltaDoctors = []
	checkedDoctorsForDelete = []
	if len(data) == 0:
		return deltaDoctors
	date = time.strftime("%m/%d/%y")
	for i in range(2,startIndex):
		a_val =  str(sheet['A'+str(i)].value)
		val = sheet['B'+str(i)].value
		val += '|'
		val += str(sheet['C'+str(i)].value)
		val += '|'
		val += str(sheet['D'+str(i)].value)
		if a_val is not None and a_val == '-':
			listOfDoctors = [l for l in listOfDoctors if l != val]
			continue
		listOfDoctors.append(val)
	for row in data:
		r = str(row[0]) + "|" + str(row[1]) + "|" + str(row[2])
		checkedDoctorsForDelete.append(r)
		if r not in listOfDoctors:
			row.append('+')
			deltaDoctors.append(row)
	for doctor in listOfDoctors:
		if doctor.split('|')[1] == data[0][1] and doctor not in checkedDoctorsForDelete:
			doctor_d = doctor
			doctor_d += '| |' + date + '| |-' 
			deltaDoctors.append(doctor_d.split('|'))
	return deltaDoctors

def readZipCodes():
	zipcodes = {}
	with open('ZipTableWithLoc.csv', 'r') as csvfile:
		reader = csv.reader(csvfile)
		for row in reader:
			val = str(row[0])
			if len(val) == 3:
				val = '00' + val
			if len(val) == 4:
				val = '0' + val
			zipcodes[val] = (row[3],row[4])
	return zipcodes

def updateDashboard(filename, wb, doctors):
	dashboard = getSheet(wb, 'Dashboard')
	ec = getEmptyCol(2, dashboard)
	dashboard[ec+'1'].value = time.strftime("%m/%d/%y")
	dashboard[ec+'2'].value = len(doctors)
	wb.save(filename)

def update_zip_info(filename,wb,data):
	ws = getSheet(wb,'Docs')
	startIndex = getStartIndex(ws)
	if ws['A'+str(1)].value is None:
		ws['A'+str(1)].value = "+/-"
		ws['B'+str(1)].value = "Doctor Name"
		ws['C'+str(1)].value = "Zipcode"
		ws['D'+str(1)].value = "Address"
		ws['E'+str(1)].value = "Date added"
		ws['F'+str(1)].value = "Contact"
	for i,row in enumerate(getChanges(ws,data,startIndex)):
		ws['A'+str(startIndex+i)].value = str(row[5])
		ws['B'+str(startIndex+i)].value = str(row[0])
		ws['C'+str(startIndex+i)].value = str(row[1])
		ws['D'+str(startIndex+i)].value = str(row[2])
		ws['E'+str(startIndex+i)].value = str(row[3])
		ws['F'+str(startIndex+i)].value = str(row[4])
	wb.save(filename)

def conformis():
	excelFile = "ConforMIS.xlsx"
	BASE_URL = "http://www.conformis.com/wp-admin/admin-ajax.php"
	wb = load_workbook(filename = excelFile,data_only=True)
	zipcodeDB = wb['Zipcodes']
	zipcodeCoordinatesMap = readZipCodes()
	doctors = []
	hospitals = []
	date = time.strftime("%m/%d/%y")
	for zipcode in get_zipcodes(zipcodeDB):
		print "Updating " + zipcode
		headers={'Content-Type':'application/x-www-form-urlencoded'}
		loc = zipcodeCoordinatesMap[zipcode]
		data = {'address':zipcode,'lat':loc[0],'lng':loc[1],'radius':25,'action':'csl_ajax_search'}
		r = requests.post(BASE_URL,headers=headers,data=data)
		office_arr = json.loads(r.text)['response']
		big_data = []
		for office in office_arr:
			if float(office['distance']) > 25:
				continue
			doc_name = office['name']
			doc_location = office['address']
			doc_practice = office['data']['identifier']
			doc_contact = remSpCh(office['phone'])
			big_data.append([doc_name,doc_practice,doc_location,date,doc_contact])
			doctors.append(doc_name)
			hospitals.append(doc_practice)
		update_zip_info(excelFile,wb,big_data)
	updateDashboard(excelFile, wb, doctors, hospitals)

conformis()
print("\n\nExecution Time: --- %2.f seconds ---\n" % (time.time() - start_time))