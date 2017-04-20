from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Font
#from geopy.geocoders import Nominatim
from geopy.distance import vincenty
import requests, time, re, json, csv
import unicodedata

'''

Author: Harshil Shah
Date Updated: 10/16/16
Project Owner: Kyle Rose

'''


def getWebData(link):
    return BeautifulSoup(requests.get(link).text,'html.parser')

def remSpCh(s):
	s = unicodedata.normalize('NFKD', s).encode('ascii','ignore')
	return ''.join([i for i in s if ord(i) < 128])

def getEmptyCol(sheet):
	for i in range(3,999):
		val = sheet[get_column_letter(i)+'2'].value
		if val is None or val == time.strftime("%m/%d/%y"):
			return get_column_letter(i)

def getTotalIndex(sheet):
	index = 1
	for i in range(1,sheet.max_row+1):
		val = sheet['A'+str(i)].value
		if val is None:
			return i
		if str(val).lower().strip() == 'total count':
			index =  i+1
	return index

def get_zipcodes(sheet):
	zipcodes = []
	for i in range(4,sheet.max_row+1):
		val = sheet['A'+str(i)].value
		if val is None:
			break
		else:
			val2 = str(sheet['B'+str(i)].value)
			if len(val2) < 5:
				val2 = '0' + val2
			zipcodes.append(val.split(',')[0] + ' - ' + val2)
	return zipcodes

def readZipCodes():
	zipcodes = {}
	with open('Zipcode_to_Coordinates.csv', 'r') as csvfile:
		reader = csv.reader(csvfile)
		for row in reader:
			zipcodes[row[0]] = (row[1],row[2])
	return zipcodes

def writeZipCoordinates(data):
	with open('Zipcode_to_Coordinates.csv', 'w') as csvfile:
		writer = csv.writer(csvfile)
		for row in data:
			writer.writerow(row)

def writeZipCodes():
	excelFile = "Medtech Web script (Intersect ENT).xlsx"
	wb = load_workbook(filename = excelFile,data_only=True)
	aggregate = wb["Aggregate"]
	zips = []
	for zipcode in get_zipcodes(aggregate):
		city = zipcode.split('-')[0].strip()
		geolocator = Nominatim()
		loc = geolocator.geocode(city)
		zips.append([zipcode,loc.latitude,loc.longitude])
	writeZipCoordinates(zips)

def addCities(cities,sheet):
	index = 0
	for i in range(4,sheet.max_row+1):
		if sheet['A'+str(i)].value is None:
			index = i
			break
	index = i
	for j in range(len(cities)):
		sheet['A'+str(index+j)].value = cities[j]


def update_zip_info(filename,sheetname,wb,data,aggregate,aggregate_row_index):
	if sheetname not in wb.get_sheet_names():
		ws = wb.create_sheet()
		ws.title = sheetname
	else:
		ws = wb[sheetname]
	startIndex = getTotalIndex(ws)
	ws['A'+str(startIndex)].value = "Date queried"
	ws['B'+str(startIndex)].value = time.strftime("%m/%d/%y")
	ws['A'+str(startIndex+1)].value = "Practice number"
	ws['B'+str(startIndex+1)].value = "Doctor Name"
	ws['C'+str(startIndex+1)].value = "Practice Name"
	ws['D'+str(startIndex+1)].value = "Location"
	ws['E'+str(startIndex+1)].value = "Contact information"
	ws['F'+str(startIndex+1)].value = "Misc information"
	itr = 2
	for i,row in enumerate(data):
		ws['A'+str(startIndex+itr)].value = itr-1
		ws['B'+str(startIndex+itr)].value = row[0]
		ws['C'+str(startIndex+itr)].value = row[1]
		ws['D'+str(startIndex+itr)].value = row[2]
		ws['E'+str(startIndex+itr)].value = row[3]
		ws['F'+str(startIndex+itr)].value = row[4]
		itr+=1
	total = itr-2
	ws['A'+str(startIndex+itr)].value = "Total count"
	ws['B'+str(startIndex+itr)].value = total
	ec = getEmptyCol(aggregate)
	aggregate[ec+'2'].value = time.strftime("%m/%d/%y")
	aggregate[ec+str(aggregate_row_index)].value = total
	wb.save(filename)

def conformis():
	excelFile = "Medtech Web script (ConforMIS).xlsx"
	BASE_URL = "http://www.conformis.com/wp-admin/admin-ajax.php"
	wb = load_workbook(filename = excelFile,data_only=True)
	aggregate = wb["Aggregate"]
	for i,sheetname in enumerate(get_zipcodes(aggregate)):
		print "Updating " + sheetname
		zipcode = sheetname.split('-')[1].strip()
		headers={'Content-Type':'application/x-www-form-urlencoded'}
		loc = readZipCodes()[sheetname]
		data = {'address':zipcode,'lat':loc[0],'lng':loc[1],'radius':200,'action':'csl_ajax_search'}
		r = requests.post(BASE_URL,headers=headers,data=data)
		office_arr = json.loads(r.text)['response']
		big_data = []
		for office in office_arr:
			if float(office['distance']) > 25:
				continue
			doc_name = office['name']
			doc_location = office['address']
			doc_practice = office['data']['identifier']
			doc_distance = office['distance'] + " mi"
			doc_contact = office['phone']
			big_data.append([doc_name,doc_practice,doc_location,doc_contact,doc_distance])
		update_zip_info(excelFile,sheetname,wb,big_data,aggregate,i+4)

def propel():
	excelFile = "Medtech Web script (Intersect ENT).xlsx"
	BASE_URL = "http://propelopens.com/locator/find-propel-physician/"
	wb = load_workbook(filename = excelFile,data_only=True)
	aggregate = wb["Aggregate"]
	for i,sheetname in enumerate(get_zipcodes(aggregate)):
		print "Updating " + sheetname
		zipcode = sheetname.split('-')[1].strip()
		url = BASE_URL + "?zipcode=" + zipcode + "&distance=25&submit=+#search"
		soup = getWebData(url)
		big_data = []
		for office in soup.find_all('div',{'class':'theoffice'}):
			doc_div = office.find('div',{'class':'names'}).span
			doc_name = doc_div.text.split(',')[0].strip()
			doc_practice = office.find('span',recursive=False).text
			doc_location = re.search(r'Address(.*)</',str(office)).group()
			doc_location = re.sub(r'\s{3,}',", ",doc_location[11:])
			doc_location = doc_location.replace('br','').replace('<','').replace('>','').replace('/','')
			doc_distance = office.div.p.strong.text.replace('Distance:','').strip()
			try:
				doc_contact = re.search(r'\d{3}-\d{3}-\d{4}',office.text).group()
			except:
				doc_contact = None
			big_data.append([doc_name,doc_practice,doc_location,doc_contact,doc_distance])
		update_zip_info(excelFile,sheetname,wb,big_data,aggregate,i+4)

def entellus():
	excelFile = "Medtech Web script (Entellus).xlsx"
	BASE_URL = "http://www.sinussurgeryoptions.com/find-a-doctor-results"
	wb = load_workbook(filename = excelFile,data_only=True)
	aggregate = wb["Aggregate"]
	for i,sheetname in enumerate(get_zipcodes(aggregate)):
		print "Updating " + sheetname
		zipcode = sheetname.split('-')[1].strip()
		url = BASE_URL + "?distance%5Bpostal_code%5D="+zipcode+"&distance%5Bsearch_distance%5D=25&distance%5Bsearch_units%5D=mile"
		soup = getWebData(url)
		big_data = []
		if soup.find('table') is not None:
			for office in soup.find('table').find_all('tr')[1:5]:
				doc_name = office.find_all('td')[0].text.strip()
				doc_practice = "Physician"
				doc_distance = office.find_all('td')[2].text.strip()
				try:
					doc_practice = office.find_all('td')[1].a.text
					office.find_all('td')[1].a.clear()
				except:
					doc_practice = None
				split_arr = office.find_all('td')[1].text.split('(')
				doc_location = split_arr[0].strip()
				if len(split_arr) < 2:
					doc_contact = None
				else:
					doc_contact = '(' + split_arr[1].strip()
				big_data.append([doc_name,doc_practice,doc_location,doc_contact,doc_distance])
		update_zip_info(excelFile,sheetname,wb,big_data,aggregate,i+4)

def zeltiq():
	excelFile = "Medtech Web script (Zeltiq).xlsx"
	BASE_URL = "http://www.sinussurgeryoptions.com/find-a-doctor-results"
	wb = load_workbook(filename = excelFile,data_only=True)
	aggregate = wb["Aggregate"]
	for i,sheetname in enumerate(get_zipcodes(aggregate)):
		big_data = []
		print "Updating " + sheetname
		city = sheetname.split('-')[0].strip()		
		loc = readZipCodes()[sheetname]
		data = {'Latitude':loc[0], 'Longitude':loc[1], 'Radius':40.23, 'SearchTerm':city}
		try:
			r = requests.post("http://find.coolsculpting.com/api/GetLocations",data=data)
		except:
			continue
		office_arr = json.loads(r.text)['locations']
		for office in office_arr:
			doc_name = office['physician1'].split(',')[0]
			doc_practice = office['name']
			doc_location = office['address1'] + " " + office['city'] + ", " + office['state'] + " " + office['zipcode']
			doc_contact = office['phone']
			doc_distance = "{:.2f} mi".format(float(office['distance']) * 0.62137)
			big_data.append([doc_name,doc_practice,doc_location,doc_contact,doc_distance])
		update_zip_info(excelFile,sheetname,wb,big_data,aggregate,i+4)

def orbera():
	excelFile = "Medtech Web script (Orbera).xlsx"
	BASE_URL = "https://www.orbera.com/apexremote"
	wb = load_workbook(filename = excelFile,data_only=True)
	aggregate = wb["Aggregate"]
	for i,sheetname in enumerate(get_zipcodes(aggregate)):
		big_data = []
		print "Updating " + sheetname
		city = sheetname.split('-')[0].strip()	
		time.sleep(1)	
		loc = readZipCodes()[sheetname]
		headers = {'Content-Type':'application/json', 'Referer':'https://www.orbera.com/find-a-specialist'}
		loc_str = str(loc[0]) + ":" + str(loc[1]) + ":mi:25:Any:undefined:abc@xyz.com"
		csrf = "VmpFPSxNakF4Tmkwd09DMHhNRlF4TkRvMU5qb3pNaTQ1TnpCYSxmLUR2dEJrRFJ2czVSUGR2WV8yNVVzLE5tVTRNamRq"
		data = "{'data': ['"+loc_str+"', '1', 'Orbera'], 'ctx': {'ns': '', 'ver': 28, 'vid': '066500000006Ka9', 'csrf': '"+csrf+"'}, 'tid': 2, 'action': 'LWFindSpecialistController', 'type': 'rpc', 'method': 'getOfficeByPage'}"
		r = requests.post(BASE_URL,headers=headers,data=data)
		office_arr = json.loads(r.text)[0]['result']
		if 'v' in office_arr:
			office_arr = office_arr['v']
		for office in office_arr:
			ofv = office['office']
			doc_name = ""
			doc_contact = ""
			if 'v' in office['surgeon']:
				if 'Name' in office['surgeon']['v']:
					doc_name = office['surgeon']['v']['Name']
			else:
				if 'Name' in office['surgeon']:
					doc_name = office['surgeon']['Name']
			if 'v' in ofv:
				ofv = ofv['v']
			doc_practice = ofv['Name']
			doc_location = ofv['Address_1__c'] + ', ' + ofv['City__c'] + ' ' + ofv['State__c'] + ' ' + ofv['Zip_Code__c']
			if 'Phone__c' in ofv:
				doc_contact = ofv['Phone__c']
			loc1 = (loc[0],loc[1])
			loc2 = (ofv['Location__Latitude__s'],ofv['Location__Longitude__s'])
			doc_distance = str(vincenty(loc1, loc2).miles) + " mi"
			big_data.append([doc_name,doc_practice,doc_location,doc_contact,doc_distance])
		update_zip_info(excelFile,sheetname,wb,big_data,aggregate,i+4)

def reshape():
	excelFile = "Medtech Web script (Reshape).xlsx"
	BASE_URL = "https://reshapeready.com/get-started/?"
	wb = load_workbook(filename = excelFile,data_only=True)
	aggregate = wb["Aggregate"]
	regex = re.compile(ur'dataFAC = {([^}]*)}')
	for i,sheetname in enumerate(get_zipcodes(aggregate)):
		big_data = []
		print "Updating " + sheetname
		zipcode = sheetname.split('-')[1].strip()
		soup = getWebData(BASE_URL+"tc_search="+zipcode+"&tc_radius=50")
		for office in re.findall(regex, soup.text):
			doc_distance = re.search("dist:\"([^\"]*)",office)
			if doc_distance:
				doc_distance = doc_distance.group(0).replace("dist:\"",'').strip()
				if int(doc_distance) > 25:
					continue
				else:
					doc_distance += " mi"
			doc_name = re.search("name:\"([^\"]*)",office)
			doc_practice = ""
			doc_location = re.search("address(.*)longitude",office)
			doc_contact = re.search("phone:\"([^\"]*)",office)
			if doc_name:
				doc_name = doc_name.group(0).replace("name:\"",'')
			if doc_location:
				doc_location = doc_location.group(0)
				for j in ['address','longitude','city','state','zip','\"',':']:
					doc_location = doc_location.replace(j,'').strip()
			if doc_contact:
				doc_contact = doc_contact.group(0).replace("phone:\"",'')
			big_data.append([doc_name,doc_practice,doc_location,doc_contact,doc_distance])
		update_zip_info(excelFile,sheetname,wb,big_data,aggregate,i+4)


def obalon():
	excelFile = "Medtech Web script (Obalon).xlsx"
	wb = load_workbook(filename = excelFile,data_only=True)
	aggregate = wb["Aggregate"]
	soup = getWebData("http://www.obalon.com/find-your-doctor/")
	cities = soup.find_all('a',{'class':'slp_directory_link'})
	cities = [x['href'].split('=')[1] for x in cities if len(x.text) > 3]
	prev_cities = [x.split('-')[0].strip() for x in get_zipcodes(aggregate)]
	cities = [x for x in cities if x not in prev_cities]
	addCities(cities,aggregate)
	wb.save(excelFile)
	if len(prev_cities) == 0:
		prev_cities = cities
	for i,sheetname in enumerate(prev_cities):
		big_data = []
		headers = {'Content-Type':'application/x-www-form-urlencoded'}
		params = "formdata=&options%5Bbubblelayout%5D=%3Cdiv+id%3D%22slp_info_bubble_%5Bslp_location+id%5D%22+class%3D%22slp_info_bubble+%5Bslp_location+featured%5D%22%3E%0D%0A%3Cstrong%3E%5Bslp_location+name+suffix+br%5D%3C%2Fstrong%3E%0D%0A%5Bslp_location+address%5D+%5Bslp_location+address2+suffix+br%5D%0D%0A%5Bslp_location+city+suffix+comma%5D+%5Bslp_location+state%5D+%5Bslp_location+zip+suffix+br%5D%0D%0A%5Bslp_location+country+suffix+br%5D%0D%0A%5Bhtml+br+ifset+directions%5D%0D%0A%5Bslp_option+label_directions+wrap+directions%5D%0D%0A%5Bhtml+br+ifset+url%5D%5Bslp_location+web_link%5D%0D%0A%5Bslp_location+email+wrap+mailto+%5D%5Bslp_option+label_email+ifset+email%5D%5Bhtml+closing_anchor+ifset+email%5D%5Bhtml+br+ifset+email%5D%0D%0A%3C%2Fdiv%3E&options%5Bignore_radius%5D=1&options%5Bmap_domain%5D=maps.google.com&options%5Bmap_end_icon%5D=http%3A%2F%2Fwww.obalon.com%2Fwp-content%2Fplugins%2Fstore-locator-le%2Fimages%2Ficons%2Fbulb_azure.png&options%5Bmap_home_icon%5D=http%3A%2F%2Fwww.obalon.com%2Fwp-content%2Fplugins%2Fstore-locator-le%2Fimages%2Ficons%2Fbulk_blue.png&options%5Bno_autozoom%5D=&options%5Bno_homeicon_at_start%5D=1&options%5Bradius_behavior%5D=always_use&options%5Bresults_layout%5D=%3Cdiv+class%3D%22col-sm-4+result-box%22%3E%0D%0A%3Cdiv+class%3D%22results_entry+location_primary+%5Bslp_location+featured%5D%22+id%3D%22slp_results_%5Bslp_location+id%5D%22%3E%0D%0A%0D%0A%3Cdiv+class%3D%22slp_results_image%22%3E%3Ca+href%3D%22%5Bslp_location+url%5D%22%3E%3Cimg+src%3D%5Bslp_location+image+type%3Dimage%5D%3E%3C%2Fa%3E%3C%2Fdiv%3E%0D%0A%3Ca+href%3D%22%5Bslp_location+url%5D%22%3E%3Cspan+class%3D%22location_name%22%3E%5Bslp_location+name+suffix+space%5D%3C%2Fspan%3E%3C%2Fa%3E%0D%0A%3Cbr%2F%3E%0D%0A%3Cdiv%3E%5Bslp_location+address%5D%0D%0A%5Bslp_location+address2+suffix+br%5D%3C%2Fdiv%3E%0D%0A%3Cdiv%3E%5Bslp_location+city_state_zip+suffix+br%5D%3C%2Fdiv%3E%0D%0A%3Cspan+class%3D%22slp_result_contact+slp_result_website%22%3E%5Bslp_location+web_link+raw%5D%3C%2Fspan%3E%0D%0A%3Cspan+class%3D%22slp_result_contact+slp_result_directions%22%3E%3Ca+href%3D%22http%3A%2F%2F%5Bslp_option+map_domain%5D%2Fmaps%3Fsaddr%3D%5Bslp_location+search_address%5D%26amp%3Bdaddr%3D%5Bslp_location+location_address%5D%22+target%3D%22_blank%22+class%3D%22storelocatorlink%22%3E%5Bslp_option+label_directions%5D%3C%2Fa%3E%3C%2Fspan%3E%0D%0A%3C%2Fdiv%3E%0D%0A%3C%2Fdiv%3E&options%5Bslplus_version%5D=4.6.3&options%5Buse_sensor%5D=0&options%5Bmessage_no_results%5D=No+locations+found.&options%5Bmessage_no_api_key%5D=This+site+most+likely+needs+a+Google+Maps+API+key.&options%5Bdistance_unit%5D=miles&options%5Bradii%5D=10%2C25%2C50%2C100%2C(200)%2C500&options%5Bmap_center%5D=United+States&options%5Bmap_center_lat%5D=37.09024&options%5Bmap_center_lng%5D=-95.712891&options%5Bzoom_level%5D=12&options%5Bzoom_tweak%5D=0&options%5Bmap_type%5D=roadmap&options%5Bimmediately_show_locations%5D=0&options%5Binitial_radius%5D=15000&options%5Binitial_results_returned%5D=25&options%5Blabel_website%5D=View+Profile&options%5Blabel_directions%5D=Get+Directions&options%5Blabel_email%5D=Email&options%5Blabel_phone%5D=Phone&options%5Blabel_fax%5D=Fax&options%5Bmap_width%5D=100&options%5Bmap_height%5D=35&options%5Blayout%5D=%3Cdiv+id%3D%22sl_div%22%3E%5Bslp_search%5D%5Bslp_map%5D%5Bslp_results%5D%3C%2Fdiv%3E&options%5Bmaplayout%5D=%5Bslp_mapcontent%5D%5Bslp_maptagline%5D&options%5Bresultslayout%5D=%3Cdiv+class%3D%22col-sm-4+result-box%22%3E%0D%0A%3Cdiv+class%3D%22results_entry+location_primary+%5Bslp_location+featured%5D%22+id%3D%22slp_results_%5Bslp_location+id%5D%22%3E%0D%0A%5Bslp_addon+section%3Dprimary+position%3Dfirst+suffix+br%5D%0D%0A%3Cdiv+class%3D%22slp_results_image%22%3E%3Ca+href%3D%22%5Bslp_location+url%5D%22%3E%3Cimg+src%3D%5Bslp_location+image+type%3Dimage%5D%3E%3C%2Fa%3E%3C%2Fdiv%3E%0D%0A%3Ca+href%3D%22%5Bslp_location+url%5D%22%3E%3Cspan+class%3D%22location_name%22%3E%5Bslp_location+name+suffix+space%5D%3C%2Fspan%3E%3C%2Fa%3E%0D%0A%3Cbr%2F%3E%0D%0A%3Cdiv%3E%5Bslp_location+address%5D%0D%0A%5Bslp_location+address2+suffix+br%5D%3C%2Fdiv%3E%0D%0A%3Cdiv%3E%5Bslp_location+city_state_zip+suffix+br%5D%3C%2Fdiv%3E%0D%0A%3Cspan+class%3D%22slp_result_contact+slp_result_website%22%3E%5Bslp_location+web_link%5D%3C%2Fspan%3E%0D%0A%3Cspan+class%3D%22slp_result_contact+slp_result_directions%22%3E%3Ca+href%3D%22http%3A%2F%2F%5Bslp_option+map_domain%5D%2Fmaps%3Fsaddr%3D%5Bslp_location+search_address%5D%26daddr%3D%5Bslp_location+location_address%5D%22+target%3D%22_blank%22+class%3D%22storelocatorlink%22%3E%5Bslp_location+directions_text%5D%3C%2Fa%3E%3C%2Fspan%3E%0D%0A%3C%2Fdiv%3E%0D%0A%3C%2Fdiv%3E&options%5Bsearchlayout%5D=%3Cdiv+id%3D%22address_search%22+class%3D%22slp+search_box%22%3E%0A++++++++%5Bslp_search_element+add_on+location%3D%22very_top%22%5D%0A++++++++%5Bslp_search_element+input_with_label%3D%22name%22%5D%0A++++++++%5Bslp_search_element+input_with_label%3D%22address%22%5D%0A++++++++%5Bslp_search_element+dropdown_with_label%3D%22city%22%5D%0A++++++++%5Bslp_search_element+dropdown_with_label%3D%22state%22%5D%0A++++++++%5Bslp_search_element+dropdown_with_label%3D%22country%22%5D%0A++++++++%5Bslp_search_element+selector_with_label%3D%22tag%22%5D%0A++++++++%5Bslp_search_element+dropdown_with_label%3D%22category%22%5D%0A++++++++%5Bslp_search_element+dropdown_with_label%3D%22gfl_form_id%22%5D%0A++++++++%5Bslp_search_element+add_on+location%3D%22before_radius_submit%22%5D%0A++++++++%3Cdiv+class%3D%22search_item%22%3E%0A++++++++++++%5Bslp_search_element+dropdown_with_label%3D%22radius%22%5D%0A++++++++++++%5Bslp_search_element+button%3D%22submit%22%5D%0A++++++++%3C%2Fdiv%3E%0A++++++++%5Bslp_search_element+add_on+location%3D%22after_radius_submit%22%5D%0A++++++++%5Bslp_search_element+add_on+location%3D%22very_bottom%22%5D%0A++++%3C%2Fdiv%3E&options%5Btheme%5D=&options%5Bid%5D=&options%5Bhide_search_form%5D=1&options%5Bcenter_map_at%5D=&options%5Bhide_map%5D=0&options%5Bshow_maptoggle%5D=0&options%5Bendicon%5D=&options%5Bhomeicon%5D=&options%5Bonly_with_category%5D=&options%5Bonly_with_tag%5D=&options%5Btags_for_pulldown%5D=&options%5Btags_for_dropdown%5D=&options%5Bhide_results%5D=0&options%5Border_by%5D=sl_distance+ASC&options%5Ballow_addy_in_url%5D=0&options%5Bhide_address_entry%5D=1&options%5Bappend_to_search%5D=&options%5Bcity%5D="+sheetname+"&options%5Bcountry%5D=&options%5Bstate%5D=&options%5Bcity_selector%5D=hidden&options%5Bcountry_selector%5D=&options%5Bstate_selector%5D=&options%5Badd_tel_to_phone%5D=1&options%5Bdisable_initial_directory%5D=0&options%5Bextended_data_version%5D=&options%5Bhide_distance%5D=1&options%5Binstalled_version%5D=4.6.3&options%5Borderby%5D=sl_distance+ASC&options%5Bshow_country%5D=1&options%5Bshow_hours%5D=1&options%5Bfeatured_location_display_type%5D=show_always&options%5Bemail_link_format%5D=popup_form&options%5Bpopup_email_title%5D=Send+An+Email&options%5Bpopup_email_from_placeholder%5D=Your+email+address.&options%5Bpopup_email_subject_placeholder%5D=Email+subject+line.&options%5Bpopup_email_message_placeholder%5D=What+do+you+want+to+say%3F&options%5Baddress_autocomplete%5D=none&options%5Baddress_autocomplete_min%5D=3&options%5Bsearchnear%5D=world&options%5Bselector_behavior%5D=use_both&options%5Bforce_load_js%5D=0&options%5Bmap_region%5D=us&options%5Bmap_options_scrollwheel%5D=false&options%5Bmap_options_scaleControl%5D=true&options%5Bmap_options_mapTypeControl%5D=true&options%5Bgoogle_map_style%5D=&options%5Bhide_bubble%5D=&options%5Bmap_initial_display%5D=map&options%5Bstarting_image%5D=&options%5Bajax_orderby_catcount%5D=0&options%5Bcron_import_timestamp%5D=&options%5Bcron_import_recurrence%5D=none&options%5Bcsv_clear_messages_on_import%5D=1&options%5Bcsv_file_url%5D=&options%5Bcsv_skip_geocoding%5D=0&options%5Bcsv_duplicates_handling%5D=update&options%5Bdefault_comments%5D=0&options%5Bdefault_icons%5D=0&options%5Bdefault_page_status%5D=publish&options%5Bdefault_trackbacks%5D=0&options%5Bload_data%5D=0&options%5Bhide_empty%5D=0&options%5Bhighlight_uncoded%5D=1&options%5Blabel_category%5D=Category%3A+&options%5Blog_import_messages%5D=1&options%5Blog_schedule_messages%5D=1&options%5Bpages_read_more_text%5D=&options%5Bpages_replace_websites%5D=1&options%5Bpage_template%5D=%5Badd_count%5D%0D%0A%3Cdiv+class%3D%22profile-header%22+id%3D%22%5Bstorepage+field%3Did%5D%22%3E%0D%0A++%3Cdiv+class%3D%22profile-schedule-consult%22%3E%0D%0A++++%3Cdiv+class%3D%22profile-headline%22%3E%3Ch2%3ESchedule+a+consultation+with+%3Cbr+%2F%3EDr.+%5Bstorepage+field%3Dstore%5D%3C%2Fh2%3E%3C%2Fdiv%3E%0D%0A++++%3Cdiv+class%3D%22profile-information%22%3E%0D%0A++++++%3Cdiv+class%3D%22profile-photo%22%3E%5Bstorepage+field%3Dimage+type%3Dimage%5D%3C%2Fdiv%3E%0D%0A++++++%3Cdiv+class%3D%22profile-contact%22%3E%0D%0A++++++++%3Cdiv+class%3D%22profile-address%22%3E%0D%0A++++++++++%3Ch3%3EDr.+%5Bstorepage+field%3Dstore%5D%3C%2Fh3%3E%3Cp%3E%3Cstrong%3E%5Bstorepage+field%3DDepartment%5D%3C%2Fstrong%3E%3C%2Fp%3E%3Cp%3E%5Bstorepage+field%3Daddress%5D%0D%0A++++++++++%5Bstorepage+field%3Daddress2%5D%0D%0A++++++++++%5Bstorepage+field%3Dcity%5D%2C+%5Bstorepage+field%3Dstate%5D++%5Bstorepage+field%3Dzip%5D%3C%2Fp%3E%0D%0A%09%09++%3Cdiv+class%3D%22profile-desk-phone%22%3E%3Ca+href%3D%22tel%3A%5Bstorepage+field%3Dphone%5D%22%3E%5Bstorepage+field%3Dphone%5D%3C%2Fa%3E%3C%2Fdiv%3E%0D%0A++++++++%3C%2Fdiv%3E%0D%0A%09%3Cdiv+title%3D%22%5Bstorepage+field%3Dstore_user%5D%22+class%3D%22docPhoneButton%22+role%3D%22%5Bstorepage+field%3D+phone%5D%22+id%3D%22agent_%5Bstorepage+field%3Dfax%5D%22%3E%5Btwilio_call_btn%5D%3C%2Fdiv%3E+%3Ca+class%3D%22profile-website-button%22+href%3D%22%5Bstorepage+field%3Durl%5D%22+type%3D%22button%22%3EVisit+Website%3C%2Fa%3E%0D%0A++++++%3C%2Fdiv%3E%0D%0A++++%3C%2Fdiv%3E%0D%0A++%3C%2Fdiv%3E%0D%0A++%3Cdiv+class%3D%22profile-request-consult%22%3E%0D%0A++++%3Cdiv+class%3D%22profile-headline%22%3E%3Ch2%3ERequest+Consultation%3C%2Fh2%3E%3C%2Fdiv%3E%0D%0A++++%3Cdiv+class%3D%22profile-form%22%3E%5Bgravityform+id%3D%222%22+title%3D%22true%22+description%3D%22false%22+ajax%3D%22true%22%5D%3C%2Fdiv%3E%0D%0A++%3C%2Fdiv%3E%0D%0A++%3Cdiv+class%3D%22profile-description%22%3E%5Bstorepage+field%3Ddescription%5D%0D%0A++%3C%2Fdiv%3E%0D%0A%3C%2Fdiv%3E&options%5Bpermalink_starts_with%5D=practices&options%5Bpermalink_flush_needed%5D=0&options%5Bprevent_new_window%5D=0&options%5Bprepend_permalink_blog%5D=1&options%5Bshow_icon_array%5D=0&options%5Bshow_legend_text%5D=0&options%5Bshow_option_all%5D=Any&options%5Bshow_cats_on_search%5D=&options%5Btag_autosubmit%5D=0&options%5Btag_dropdown_first_entry%5D=&options%5Btag_label%5D=&options%5Btag_selector%5D=dropdown&options%5Btag_selections%5D=&options%5Btag_show_any%5D=1&options%5Btag_output_processing%5D=replace_with_br&options%5Bterritory%5D=&options%5Bterritory_selector%5D=&options%5Breporting_enabled%5D=1&options%5Buse_contact_fields%5D=1&options%5Buse_nonces%5D=0&options%5Buse_pages%5D=1&options%5Bcsv_first_line_has_field_name%5D=1&options%5Bcsv_skip_first_line%5D=0&options%5Bcustom_css%5D=&options%5Buse_same_window%5D=0&radius=15000&tags=&action=csl_ajax_onload"
		print "Updating " + sheetname
		r = requests.post("http://www.obalon.com/wp-admin/admin-ajax.php",headers=headers,data=params)
		office_arr = json.loads(r.text)["response"]
		for office in office_arr:
			doc_name = office['data']['sl_store']
			try:
				doc_practice = office['department']
			except:
				doc_practice = ""
			doc_location = office['address'] + ' ' + office['city'] + ', ' + office['state'] + ' ' + office['zip']
			doc_contact = office['data']['sl_phone']
			big_data.append([doc_name,doc_practice,doc_location,doc_contact,''])
		update_zip_info(excelFile,sheetname,wb,big_data,aggregate,i+4)




start_time = time.time()

'''
print "\nScraping Company 1 (ConforMIS):"
conformis()
print "\n\nScraping Company 2 (Intersect ENT):"
propel()
print "\n\nScraping Company 3 (Entellus):"
entellus()
print "\n\nScraping Company 4 (Zeltiq):"
zeltiq()
print "\n\nScraping Company 5 (Reshape):"
reshape()
print "\n\nScraping Company 6 (Orbera):"
orbera()
'''
print "\n\nScraping Company 7 (Obalon):"
obalon()


print("\n\nExecution Time: --- %2.f seconds ---\n" % (time.time() - start_time))