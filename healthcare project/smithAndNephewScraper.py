from bs4 import BeautifulSoup
from openpyxl import load_workbook
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from geopy.distance import vincenty
import requests, time, re, json, csv
import unicodedata

start_time = time.time()
date = time.strftime("%m/%d/%y")
baseProcedureToColMapStryker = { "Knee": 'G', "Partial Knee": 'H', "Robotic-arm assisted partial knee": 'I', 
		"Robotic-Arm Assisted Total Knee":"J", "Hip": 'K', "Robotic-Arm Assisted Hip": 'L', "Mobile Bearing Hip": 'M', 
		"Direct Anterior Approach": 'N', "Partial Knee Replacement": 'O', "Ankle": 'P', "STAR ankle replacement": 'Q', 
		"Ankle Arthritis": 'R', "Foot": 'S', "Foot Arthritis": 'T', "Flatfoot": 'U', "Bunion": 'V', "Hammertoe": 'W', 
		"Charcot": 'X', "Sports Injuries": 'Y'}

def getWebData(link):
    return BeautifulSoup(requests.get(link).text,'html.parser')

def remSpCh(s):
	if s is None:
		return ''
	s = unicodedata.normalize('NFKD', s).encode('ascii','ignore')
	return ''.join([i for i in s if ord(i) < 128])

def soupTextIfNotNone(obj):
	if obj is not None:
		return remSpCh(obj.text)
	return remSpCh(obj)

def getSheet(wb,sheetname):
	if sheetname in wb.get_sheet_names():
		return wb[sheetname]
	ws = wb.create_sheet()
	ws.title = sheetname
	return ws

def getEmptyCol(start_col, sheet):
	for i in range(start_col,999):
		val = sheet[get_column_letter(i)+'1'].value
		if val is None or val == time.strftime("%m/%d/%y"):
			return get_column_letter(i)

def getStartIndex(sheet):
	for i in range(2,sheet.max_row+2):
		val = sheet['B'+str(i)].value
		if val is None:
			return i
	return i

def get_zipcodes(sheet):
	zipcodes = []
	for i in range(1,sheet.max_row+1):
		val = sheet['A'+str(i)].value
		if val is None:
			break
		else:
			val = str(val)
			if len(val.split('-')[0].strip()) < 5:
				val = '0' + val
			zipcodes.append(val)
	return zipcodes

def getChanges(sheet,data,startIndex):
	listOfDoctors = []
	deltaDoctors = []
	checkedDoctorsForDelete = []
	if len(data) == 0:
		return deltaDoctors
	date = time.strftime("%m/%d/%y")
	for i in range(2,startIndex):
		a_val =  str(sheet['A'+str(i)].value)
		val = sheet['B'+str(i)].value
		val += '|'
		val += str(sheet['C'+str(i)].value)
		val += '|'
		val += str(sheet['D'+str(i)].value)
		if a_val is not None and a_val == '-':
			listOfDoctors = [l for l in listOfDoctors if l != val]
			continue
		listOfDoctors.append(val)
	for row in data:
		r = str(row[0]) + "|" + str(row[1]) + "|" + str(row[2])
		checkedDoctorsForDelete.append(r)
		if r not in listOfDoctors:
			row.append('+')
			deltaDoctors.append(row)
	for doctor in listOfDoctors:
		if doctor.split('|')[2] == data[0][2] and doctor not in checkedDoctorsForDelete:
			doctor_d = doctor
			doctor_d += '| |' + date + '|-' 
			deltaDoctors.append(doctor_d.split('|'))
	return deltaDoctors

def updateDashboard(filename, wb, doctors, hospitals):
	dashboard = getSheet(wb, 'Dashboard')
	dashboard['A2'].value = 'Discrete doctors'
	dashboard['A3'].value = 'Discrete hospitals'
	ec = getEmptyCol(2, dashboard)
	dashboard[ec+'1'].value = time.strftime("%m/%d/%y")
	dashboard[ec+'2'].value = len(set(doctors))
	dashboard[ec+'3'].value = len(set(hospitals))
	wb.save(filename)

def update_zip_info_stryker_new(filename,wb,data):
	ws = getSheet(wb,'NAVIO Docs')
	startIndex = getStartIndex(ws)
	if ws['A'+str(1)].value is None:
		ws['A'+str(1)].value = "+/-"
		ws['B'+str(1)].value = "Doctor Name"
		ws['C'+str(1)].value = "Practice Name"
		ws['D'+str(1)].value = "Zipcode - City"
		ws['E'+str(1)].value = "Address"
		ws['F'+str(1)].value = "Date added"
	for i,row in enumerate(getChanges(ws,data,startIndex)):
		ws['A'+str(startIndex+i)].value = str(row[5])
		ws['B'+str(startIndex+i)].value = str(row[0])
		ws['C'+str(startIndex+i)].value = str(row[1])
		ws['D'+str(startIndex+i)].value = str(row[2])
		ws['E'+str(startIndex+i)].value = str(row[3])
		ws['F'+str(startIndex+i)].value = str(row[4])
	wb.save(filename)
	return 0

def smithAndNephew():
	BASE_URL = "http://www.rediscoveryourgo.com/findadoctor.aspx?zipcode="
	excelFile = "Medtech Web script (Smith&Nephew).xlsx"
	wb = load_workbook(filename = excelFile,data_only=True)
	zipcodeDB = wb['Zipcodes']
	doctors = []
	hospitals = []
	date = time.strftime("%m/%d/%y")
	for sheetname in get_zipcodes(zipcodeDB)[:]:
		big_data = []
		names = []
		zipcode = sheetname.split('-')[1].strip()
		print "Scraping data for " + sheetname
		soup = getWebData(BASE_URL+zipcode+'&product=36')
		for listing in soup.find_all('div',{'class':'listing-container'}):
			doc_name = remSpCh(listing.find('h2').text.strip())
			if doc_name in names:
				continue
			names.append(doc_name)
			doc_practice = remSpCh(listing.find('p').text.split('  ')[0].strip())
			try:
				doc_location = ','.join(listing.find('p').text.split('        ')[1:5]).replace(',',' ').strip()
			except:
				doc_location = ''
			big_data.append([doc_name,doc_practice,sheetname,doc_location,date])
			doctors.append(doc_name)
			hospitals.append(doc_practice)
		update_zip_info_stryker_new(excelFile,wb,big_data)
	updateDashboard(excelFile,wb,doctors,hospitals)

smithAndNephew()
print("\n\nExecution Time: --- %2.f seconds ---\n" % (time.time() - start_time))