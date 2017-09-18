from bs4 import BeautifulSoup
from openpyxl import load_workbook
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from geopy.distance import vincenty
import requests, time, re, json, csv
import unicodedata

start_time = time.time()

def getWebData(link):
    return BeautifulSoup(requests.get(link).text,'html.parser')

def remSpCh(s):
	if s is None:
		return ''
	s = unicodedata.normalize('NFKD', s).encode('ascii','ignore')
	return ''.join([i for i in s if ord(i) < 128])

def getSheet(wb,sheetname):
	if sheetname in wb.get_sheet_names():
		return wb[sheetname]
	ws = wb.create_sheet()
	ws.title = sheetname
	return ws

def getEmptyCol(start_col, sheet):
	for i in range(start_col,999):
		val = sheet[get_column_letter(i)+'1'].value
		if val is None or val == time.strftime("%m/%d/%y"):
			return get_column_letter(i)

def getStartIndex(sheet):
	for i in range(2,sheet.max_row+2):
		val = sheet['B'+str(i)].value
		if val is None:
			return i
	return i

def get_zipcodes(sheet):
	zipcodes = []
	for i in range(1,sheet.max_row+1):
		val = sheet['A'+str(i)].value
		if val is None:
			break
		else:
			val = str(val)
			if len(val.split('-')[0].strip()) < 5:
				val = '0' + val
			zipcodes.append(val)
	return zipcodes

def getChanges(sheet,data,startIndex):
	listOfDoctors = []
	deltaDoctors = []
	checkedDoctorsForDelete = []
	if len(data) == 0:
		return deltaDoctors
	date = time.strftime("%m/%d/%y")
	for i in range(3,startIndex):
		a_val =  str(sheet['A'+str(i)].value)
		val = sheet['B'+str(i)].value
		val += '|'
		val += str(sheet['C'+str(i)].value)
		val += '|'
		val += str(sheet['D'+str(i)].value)
		if a_val is not None and a_val == '-':
			listOfDoctors = [l for l in listOfDoctors if l != val]
			continue
		listOfDoctors.append(val)
	for row in data:
		r = str(row[0]) + "|" + str(row[1]) + "|" + str(row[2])
		checkedDoctorsForDelete.append(r)
		if r not in listOfDoctors:
			row.append('+')
			deltaDoctors.append(row)
	for doctor in listOfDoctors:
		if doctor.split('|')[2] == data[0][2] and doctor not in checkedDoctorsForDelete:
			doctor_d = doctor
			doctor_d += '| |' + date + '|-' 
			deltaDoctors.append(doctor_d.split('|'))
	return deltaDoctors

def updateDashboard(filename, wb, doctors, hospitals):
	dashboard = getSheet(wb, 'Dashboard')
	dashboard['A2'].value = 'Discrete doctors'
	dashboard['A3'].value = 'Discrete hospitals'
	ec = getEmptyCol(2, dashboard)
	dashboard[ec+'1'].value = time.strftime("%m/%d/%y")
	dashboard[ec+'2'].value = len(set(doctors))
	dashboard[ec+'3'].value = len(set(hospitals))
	wb.save(filename)

def update_zip_info_new(filename,wb,data):
	ws = getSheet(wb,'Docs')
	startIndex = getStartIndex(ws)
	ws['A'+str(1)].value = "+/-"
	ws['B'+str(1)].value = "Doctor Name"
	ws['C'+str(1)].value = "Zipcode"
	ws['D'+str(1)].value = "Address"
	ws['E'+str(1)].value = "Date Added"
	ws['F'+str(1)].value = "Contact"
	for i,row in enumerate(getChanges(ws,data,startIndex)):
		ws['A'+str(startIndex+i)].value = row[5]
		ws['B'+str(startIndex+i)].value = row[0]
		ws['C'+str(startIndex+i)].value = row[2]
		ws['D'+str(startIndex+i)].value = row[1]
		ws['E'+str(startIndex+i)].value = row[4]
		ws['F'+str(startIndex+i)].value = row[3]
	wb.save(filename)
	return 0

def miradry():
	BASE_URL = "https://miradry.com/locate/?loc="
	excelFile = "Miramar.xlsx"
	wb = load_workbook(filename = excelFile, data_only=True)
	zipcodeDB = wb['Zipcodes']
	doctors = []
	hospitals = []
	date = time.strftime("%m/%d/%y")
	for sheetname in get_zipcodes(zipcodeDB):
		big_data = []
		print "Updating " + sheetname
		soup = getWebData(BASE_URL+sheetname+'&distance=50')
		office_arr = soup.find_all('div',{'class':'columns medium-8 info-box'})
		for office in office_arr:
			try:
				doc_name = remSpCh(office.find('h3',{'class':'dr-title'}).text)
			except:
				print "Skipping a doctor, because no doctor name found."
				continue
			try:
				doc_location = remSpCh(office.find('p',{'class':'address'}).text)
			except:
				doc_location = ''
			try:
				doc_contact = remSpCh(office.find('p',{'class':'phone'}).text)
			except:
				doc_location = ''
			doctors.append(doc_name)
			hospitals.append(doc_location)
			big_data.append([doc_name,doc_location,sheetname,doc_contact,date])
		update_zip_info_new(excelFile,wb,big_data)
	updateDashboard(excelFile, wb, doctors, hospitals)

miradry()
print("\n\nExecution Time: --- %2.f seconds ---\n" % (time.time() - start_time))