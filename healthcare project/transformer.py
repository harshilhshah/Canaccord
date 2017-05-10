from openpyxl import load_workbook
from openpyxl import Workbook
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import time, unicodedata
from shutil import copyfile

'''

Author: Harshil Shah
Date Updated: 3/28/17
Project Owner: Kyle Rose

'''

start_time = time.time()

def remSpCh(s):
	if s is None:
		return None
	try:
		s = unicodedata.normalize('NFKD', s).encode('ascii','ignore')
	except:
		return s
	return ''.join([i for i in s if ord(i) < 128])

def isDate(s):
	if len(s) < 8:
		return False
	return s[2] == '/' and s[5] == '/'

def createBackup(wb,filename):
	backup_wb = Workbook()
	backup_wb.save("Medtech Web script (" + filename + ") backup.xlsx")
	copyfile("Medtech Web script ("+filename+").xlsx","Medtech Web script (" + filename + ") backup.xlsx")

def getDeletes(oldList, newList):
	deletes = []
	for item in oldList:
		if item not in newList:
			deletes.append(item)
	return deletes

def numToLetter(num):
	if num == 0:
		return 'A'
	elif num == 1:
		return 'B'
	elif num == 2:
		return 'C'
	elif num == 3:
		return 'D'
	elif num == 4:
		return 'E'
	elif num == 5:
		return 'F'

def convertOldWayToNewWay(wb,filename):
	createBackup(wb, filename)
	for sheetname in wb.get_sheet_names():
		if sheetname == 'Aggregate':
			continue
		print "Working with " + sheetname
		ws = wb[sheetname]
		new_data = []
		unique_doctors = []
		first = True
		totalAdds = 0
		new_doctors = []
		for x in range(1,ws.max_row+1):
			val = remSpCh(ws['B'+str(x)].value)
			if val is None:
				break
			row = []
			row.append(remSpCh(ws['A'+str(x)].value))
			row.append(remSpCh(ws['B'+str(x)].value))
			row.append(remSpCh(ws['C'+str(x)].value))
			row.append(remSpCh(ws['D'+str(x)].value))
			row.append(remSpCh(ws['E'+str(x)].value))
			row.append(remSpCh(ws['F'+str(x)].value))
			new_data.append(row)
			if val == 'Doctor Name':
				new_data[-1][0] = 'Added/Deleted'
			if type(val) != str:
				deletes = getDeletes(unique_doctors,new_doctors)
				new_data[-1][0] = 'Total Changes'
				new_data[-1][1] = totalAdds + len(deletes)
				last_row = new_data.pop(-1)
				for delete in deletes:
					unique_doctors.remove(delete)
					split_delete = delete.split('|')
					split_delete.insert(0,'-')
					new_data.append(split_delete)
				new_data.append(last_row)
				totalAdds = 0
				new_doctors = []
				continue
			if isDate(val) and x != 1 and len(unique_doctors) != 0:
				first = False
			if type(val) == str and not isDate(val) and val != 'Doctor Name':
				if ws['C'+str(x)].value is not None:
					val += '|' + remSpCh(ws['C'+str(x)].value)
				if ws['D'+str(x)].value is not None:
					val += '|' + remSpCh(ws['D'+str(x)].value)
				new_doctors.append(val)
				if first:
					new_data[-1][0] = '+'
					totalAdds += 1
					unique_doctors.append(val)
				else:
					if val in unique_doctors:
						del new_data[-1]
					else:
						new_data[-1][0] = '+'
						totalAdds += 1
						unique_doctors.append(val)
		wb.remove_sheet(ws)
		ws = wb.create_sheet()
		ws.title = sheetname
		for rowItrNum, excelRow in enumerate(new_data):
			for colItr, cell in enumerate(excelRow):
				ws[numToLetter(colItr)+str(rowItrNum+1)].value = cell
		wb.save("Medtech Web script ("+filename+").xlsx")


def conformis():
	excelFile = "Medtech Web script (ConforMIS).xlsx"
	wb = load_workbook(filename = excelFile,data_only=True)
	convertOldWayToNewWay(wb,"ConforMIS")

def propel():
	excelFile = "Medtech Web script (Intersect ENT).xlsx"
	wb = load_workbook(filename = excelFile,data_only=True)
	convertOldWayToNewWay(wb,"Intersect ENT")

def entellus():
	excelFile = "Medtech Web script (Entellus).xlsx"
	wb = load_workbook(filename = excelFile,data_only=True)
	convertOldWayToNewWay(wb,"Entellus")

def zeltiq():
	excelFile = "Medtech Web script (Zeltiq).xlsx"
	wb = load_workbook(filename = excelFile,data_only=True)
	convertOldWayToNewWay(wb,"Zeltiq")

def reshape():
	excelFile = "Medtech Web script (Reshape).xlsx"
	wb = load_workbook(filename = excelFile,data_only=True)
	convertOldWayToNewWay(wb,"Reshape")

def orbera():
	excelFile = "Medtech Web script (Orbera).xlsx"
	wb = load_workbook(filename = excelFile,data_only=True)
	convertOldWayToNewWay(wb,"Orbera")

def obalon():
	excelFile = "Medtech Web script (Obalon).xlsx"
	wb = load_workbook(filename = excelFile,data_only=True)
	convertOldWayToNewWay(wb,"Obalon")

def stryker():
	excelFile = "Medtech Web script (Stryker).xlsx"
	wb = load_workbook(filename = excelFile,data_only=True)
	convertOldWayToNewWay(wb,"Stryker")

def smithAndNephew():
	excelFile = "Medtech Web script (SmithAndNephew).xlsx"
	wb = load_workbook(filename = excelFile,data_only=True)
	convertOldWayToNewWay(wb,"SmithAndNephew")

print "\nConverting Company 1 (ConforMIS):"
#conformis()
print "\n\nConverting Company 2 (Intersect ENT):"
#propel()
print "\n\nConverting Company 3 (Entellus):"
#entellus()
print "\n\nConverting Company 4 (Zeltiq):"
#zeltiq()
print "\n\nConverting Company 5 (Reshape):"
#reshape()
print "\n\nConverting Company 6 (Orbera):"
orbera()
print "\n\nConverting Company 7 (Obalon):"
obalon()
print "\n\nConverting Company 8 (Stryker):"
stryker()
print "\n\nConverting Company 9 (SmithAndNephew):"
smithAndNephew()