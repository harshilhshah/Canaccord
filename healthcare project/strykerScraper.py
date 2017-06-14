from bs4 import BeautifulSoup
from openpyxl import load_workbook
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from geopy.distance import vincenty
import requests, time, re, json, csv
import unicodedata

start_time = time.time()
date = time.strftime("%m/%d/%y")
baseProcedureToColMapStryker = { "Knee": 'G', "Partial Knee": 'H', "Robotic-arm assisted partial knee": 'I', 
		"Robotic-Arm Assisted Total Knee":"J", "Hip": 'K', "Robotic-Arm Assisted Hip": 'L', "Mobile Bearing Hip": 'M', 
		"Direct Anterior Approach": 'N', "Partial Knee Replacement": 'O', "Ankle": 'P', "STAR ankle replacement": 'Q', 
		"Ankle Arthritis": 'R', "Foot": 'S', "Foot Arthritis": 'T', "Flatfoot": 'U', "Bunion": 'V', "Hammertoe": 'W', 
		"Charcot": 'X', "Sports Injuries": 'Y'}

def getWebData(link):
    return BeautifulSoup(requests.get(link).text,'html.parser')

def remSpCh(s):
	if s is None:
		return ''
	s = unicodedata.normalize('NFKD', s).encode('ascii','ignore')
	return ''.join([i for i in s if ord(i) < 128])

def soupTextIfNotNone(obj):
	if obj is not None:
		return remSpCh(obj.text)
	return remSpCh(obj)

def getSheet(wb,sheetname):
	if sheetname in wb.get_sheet_names():
		return wb[sheetname]
	ws = wb.create_sheet()
	ws.title = sheetname
	return ws

def getEmptyCol(start_col, sheet):
	for i in range(start_col,999):
		val = sheet[get_column_letter(i)+'1'].value
		if val is None or val == time.strftime("%m/%d/%y"):
			return get_column_letter(i)

def getStartIndex(sheet):
	for i in range(2,sheet.max_row+2):
		val = sheet['B'+str(i)].value
		if val is None:
			return i
	return i

def get_zipcodes(sheet):
	zipcodes = []
	for i in range(1,sheet.max_row+1):
		val = sheet['A'+str(i)].value
		if val is None:
			break
		else:
			val = str(val)
			if len(val.split('-')[0].strip()) < 5:
				val = '0' + val
			zipcodes.append(val)
	return zipcodes

def getChanges(sheet,data,startIndex):
	listOfDoctors = []
	deltaDoctors = []
	checkedDoctorsForDelete = []
	if len(data) == 0:
		return deltaDoctors
	date = time.strftime("%m/%d/%y")
	for i in range(2,startIndex):
		a_val =  str(sheet['A'+str(i)].value)
		val = sheet['B'+str(i)].value
		val += '|'
		val += str(sheet['C'+str(i)].value)
		val += '|'
		val += str(sheet['D'+str(i)].value)
		if a_val is not None and a_val == '-':
			listOfDoctors = [l for l in listOfDoctors if l != val]
			continue
		listOfDoctors.append(val)
	for row in data:
		r = str(row[0]) + "|" + str(row[1]) + "|" + str(row[2])
		checkedDoctorsForDelete.append(r)
		if r not in listOfDoctors:
			row.append('+')
			deltaDoctors.append(row)
	for doctor in listOfDoctors:
		if doctor.split('|')[2] == data[0][2] and doctor not in checkedDoctorsForDelete:
			doctor_d = doctor
			doctor_d += '| |' + date + '|-' 
			deltaDoctors.append(doctor_d.split('|'))
	return deltaDoctors

def getUniqueCountForProcedure(wb,dashboard,procedure):
	ws = getSheet(wb, 'Docs')
	procedureMap = {}
	letter_index =  7
	while ws[get_column_letter(letter_index)+str(1)].value is not None:
		procedureMap[ws[get_column_letter(letter_index)+str(1)].value.upper()] = get_column_letter(letter_index)
		letter_index += 1
	uniqueList = []
	for i in range(2,ws.max_row+2):
		val = ws[procedureMap[procedure.upper()]+str(i)].value
		if val is not None and val == 'x':
			uniqueList.append(ws['B'+str(i)].value+"|"+ws['C'+str(i)].value)
	return len(set(uniqueList))

def updateDashboard(filename, wb):
	dashboard = getSheet(wb, 'Dashboard')
	ec = getEmptyCol(2, dashboard)
	dashboard[ec+'1'].value = time.strftime("%m/%d/%y")
	itr = 2
	while dashboard['A'+str(itr)].value is not None:
		procedure = dashboard['A'+str(itr)].value 
		dashboard[ec+str(itr)].value = getUniqueCountForProcedure(wb,dashboard,procedure)
		itr += 1
	while dashboard['A'+str(itr)].value != 'Hospital':
		itr += 1
	itr += 1
	while dashboard['A'+str(itr)].value is not None:
		procedure = dashboard['A'+str(itr)].value 
		dashboard[ec+str(itr)].value = getUniqueCountForProcedure(wb,dashboard,procedure)
		itr += 1
	wb.save(filename)

def update_zip_info_stryker_new(filename,wb,data):
	ws = getSheet(wb,'Docs')
	startIndex = getStartIndex(ws)
	if ws['A'+str(1)].value is None:
		ws['A'+str(1)].value = "+/-"
		ws['B'+str(1)].value = "Doctor Name"
		ws['C'+str(1)].value = "Practice Name"
		ws['D'+str(1)].value = "Zipcode - City"
		ws['E'+str(1)].value = "Address"
		ws['F'+str(1)].value = "Date added"
		for item in sorted(baseProcedureToColMapStryker.items(), key=lambda (key,value): value):
			ws[item[1]+str(startIndex+1)].value = item[0]
	procedureMap = {}
	letter_index =  7
	while ws[get_column_letter(letter_index)+str(1)].value is not None:
		procedureMap[ws[get_column_letter(letter_index)+str(1)].value.upper()] = get_column_letter(letter_index)
		letter_index += 1
	for i,row in enumerate(getChanges(ws,data,startIndex)):
		ws['A'+str(startIndex+i)].value = str(row[6])
		ws['B'+str(startIndex+i)].value = str(row[0])
		ws['C'+str(startIndex+i)].value = str(row[1])
		ws['D'+str(startIndex+i)].value = str(row[2])
		ws['E'+str(startIndex+i)].value = str(row[3])
		ws['F'+str(startIndex+i)].value = str(row[4])
		for procedure in row[5]:
			try:	
				ws[procedureMap[procedure.upper()]+str(startIndex+i)].value = 'x'
			except:
				print "Error: There's no such column: " + procedure + ". Please add a column with this name"
				continue
	wb.save(filename)
	return 0

def stryker():
	BASE_URL = "https://patients.stryker.com/surgeons/?distance=100&techs%5Bknee%5D=1&techs%5Bhip%5D=1&search="
	excelFile = "Stryker.xlsx"
	wb = load_workbook(filename = excelFile,data_only=True)
	zipcodeDB = wb['Zipcodes']
	doctors = []
	hospitals = []
	date = time.strftime("%m/%d/%y")
	for sheetname in get_zipcodes(zipcodeDB):
		big_data = []
		j = 1
		zipcode = sheetname.split('-')[1].strip()
		print "Scraping data for " + sheetname
		while True:
			soup = getWebData(BASE_URL+zipcode+'&page='+str(j))
			if soup.find('div',{'class':'listing'}) is None:
				break
			for row in soup.find_all('div',{'class':'listing'}):
				doc_name = soupTextIfNotNone(row.find('span',{'class':'surgeonName'}))
				doc_practice = soupTextIfNotNone(row.find('div',{'class':'practiceName'}))
				doc_location = soupTextIfNotNone(row.find('li',{'class':'surgeonAddress'}))
				doc_procedures = [x.text for x in row.find_all('span',{'class':'badge'})]
				big_data.append([doc_name,doc_practice,sheetname,doc_location,date,doc_procedures])
				doctors.append(doc_name)
				hospitals.append(doc_practice)
			j += 1
		update_zip_info_stryker_new(excelFile,wb,big_data)
	updateDashboard(excelFile, wb)

stryker()
print("\n\nExecution Time: --- %2.f seconds ---\n" % (time.time() - start_time))