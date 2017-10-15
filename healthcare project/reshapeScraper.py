from bs4 import BeautifulSoup
from openpyxl import load_workbook
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from geopy.distance import vincenty
import requests, time, re, json, csv
import unicodedata

start_time = time.time()
date = time.strftime("%m/%d/%y")

def getWebData(link):
    return BeautifulSoup(requests.get(link).text,'html.parser')

def remSpCh(s):
	if s is None:
		return ''
	s = unicodedata.normalize('NFKD', s).encode('ascii','ignore')
	return ''.join([i for i in s if ord(i) < 128])

def soupTextIfNotNone(obj):
	if obj is not None:
		return remSpCh(obj.text)
	return remSpCh(obj)

def getSheet(wb,sheetname):
	if sheetname in wb.get_sheet_names():
		return wb[sheetname]
	ws = wb.create_sheet()
	ws.title = sheetname
	return ws

def getEmptyCol(start_col, sheet):
	for i in range(start_col,999):
		val = sheet[get_column_letter(i)+'1'].value
		if val is None or val == time.strftime("%m/%d/%y"):
			return get_column_letter(i)

def getStartIndex(sheet):
	for i in range(2,sheet.max_row+2):
		val = sheet['B'+str(i)].value
		if val is None:
			return i
	return i

def get_zipcodes(sheet):
	zipcodes = []
	for i in range(1,sheet.max_row+1):
		val = sheet['A'+str(i)].value
		if val is None:
			break
		else:
			val = str(val)
			if len(val.split('-')[0].strip()) < 5:
				val = '0' + val
			zipcodes.append(val)
	return zipcodes

def readZipCodes():
	zipcodes = {}
	with open('ZipTableWithLoc.csv', 'r') as csvfile:
		reader = csv.reader(csvfile)
		for row in reader:
			val = str(row[0])
			if len(val) == 3:
				val = '00' + val
			if len(val) == 4:
				val = '0' + val
			zipcodes[val] = (row[3],row[4])
	return zipcodes

def getChanges(sheet,data,startIndex):
	listOfDoctors = []
	deltaDoctors = []
	checkedDoctorsForDelete = []
	if len(data) == 0:
		return deltaDoctors
	date = time.strftime("%m/%d/%y")
	for i in range(2,startIndex):
		a_val =  str(sheet['A'+str(i)].value)
		val = sheet['B'+str(i)].value
		val += '|'
		val += str(sheet['C'+str(i)].value)
		val += '|'
		val += str(sheet['D'+str(i)].value)
		if a_val is not None and a_val == '-':
			listOfDoctors = [l for l in listOfDoctors if l != val]
			continue
		listOfDoctors.append(val)
	for row in data:
		r = str(row[0]) + "|" + str(row[1]) + "|" + str(row[2])
		checkedDoctorsForDelete.append(r)
		if r not in listOfDoctors:
			row.append('+')
			deltaDoctors.append(row)
	for doctor in listOfDoctors:
		if doctor.split('|')[1] == data[0][1] and doctor not in checkedDoctorsForDelete:
			doctor_d = doctor
			doctor_d += '| |' + date + '| |-' 
			deltaDoctors.append(doctor_d.split('|'))
	return deltaDoctors

def updateDashboard(filename, wb, doctors):
	dashboard = getSheet(wb, 'Dashboard')
	ec = getEmptyCol(2, dashboard)
	dashboard[ec+'1'].value = time.strftime("%m/%d/%y")
	dashboard[ec+'2'].value = len(doctors)
	wb.save(filename)

def update_zip_info(filename,wb,data):
	ws = getSheet(wb,'Docs')
	startIndex = getStartIndex(ws)
	if ws['A'+str(1)].value is None:
		ws['A'+str(1)].value = "+/-"
		ws['B'+str(1)].value = "Doctor Name"
		ws['C'+str(1)].value = "Zipcode"
		ws['D'+str(1)].value = "Address"
		ws['E'+str(1)].value = "Date added"
		ws['F'+str(1)].value = "Contact"
	for i,row in enumerate(getChanges(ws,data,startIndex)):
		ws['A'+str(startIndex+i)].value = str(row[5])
		ws['B'+str(startIndex+i)].value = str(row[0])
		ws['C'+str(startIndex+i)].value = str(row[1])
		ws['D'+str(startIndex+i)].value = str(row[2])
		ws['E'+str(startIndex+i)].value = str(row[3])
		ws['F'+str(startIndex+i)].value = str(row[4])
	wb.save(filename)

def reshape():
	excelFile = "Reshape.xlsx"
	BASE_URL = "https://reshapeready.com/get-started/?"
	wb = load_workbook(filename = excelFile,data_only=True)
	regex = re.compile(ur'dataFAC = {([^}]*)}')
	zipcodeDB = wb['Zipcodes']
	doctors = []
	hospitals = []
	date = time.strftime("%m/%d/%y")
	for zipcode in get_zipcodes(zipcodeDB):
		big_data = []
		print "Updating " + zipcode
		soup = getWebData(BASE_URL+"tc_search="+zipcode+"&tc_radius=50")
		for office in re.findall(regex, soup.text):
			doc_distance = re.search("dist:\"([^\"]*)",office)
			if doc_distance:
				doc_distance = doc_distance.group(0).replace("dist:\"",'').strip()
				if int(doc_distance) > 25:
					continue
			doc_name = re.search("name:\"([^\"]*)",office)
			doc_location = re.search("address(.*)longitude",office)
			doc_contact = re.search("phone:\"([^\"]*)",office)
			if doc_name:
				doc_name = remSpCh(doc_name.group(0).replace("name:\"",''))
			if doc_location:
				doc_location = doc_location.group(0)
				for j in ['address','longitude','city','state','zip','\"',':']:
					doc_location = doc_location.replace(j,'').strip()
			if doc_contact:
				doc_contact = remSpCh(doc_contact.group(0).replace("phone:\"",''))
			big_data.append([doc_name,zipcode,doc_location,date,doc_contact])
			doctors.append(doc_name)
			hospitals.append(doc_practice)
		update_zip_info(excelFile,wb,big_data)
	updateDashboard(excelFile, wb, doctors, hospitals)

reshape()
print("\n\nExecution Time: --- %2.f seconds ---\n" % (time.time() - start_time))