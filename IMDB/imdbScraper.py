from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Font
import requests, time, re
import unicodedata

### Assumptions ###
# - Sheet1 won't be renamed

# Global Variables
start_time = time.time()
# Name of the file where the data gets saved.
excelFile = "IMDB.xlsx"
sheetName = 'Sheet1'
BASE_URL = "http://www.imdb.com"
wb = load_workbook(filename = excelFile,data_only=True)

def getWebData(link):
    return BeautifulSoup(requests.get(link).text,'html.parser')

def remSpCh(s):
	if type(s) == long:
		return str(s)
	s = unicodedata.normalize('NFKD', s).encode('ascii','ignore')
	return ''.join([i for i in s if ord(i) < 128])

def remWS(s):
	return ' '.join(remSpCh(s).split())

def getColumnMap(sheet):
	column_mapping = {}
	i = 0
	val = ''
	while i < 5 and val.lower() != 'title':
		i += 1
		val = sheet['A'+str(i)].value
		if val is None:
			val = ''
	for j in range(2,15):
		val = sheet[get_column_letter(j)+str(i)].value
		if val is not None:
			column_mapping[remSpCh(val).split()[0].strip().lower()] = get_column_letter(j)
	return column_mapping

def getTitles(sheet):
	titles_mapping = {}
	i = 0
	val = ''
	while i < 5 and val.lower() != 'title':
		i += 1
		val = sheet['A'+str(i)].value
		if val is None:
			val = ''
	for j in range(i+1,sheet.max_row+1):
		val = sheet['A'+str(j)].value
		if val is not None:
			titles_mapping[remSpCh(val).strip()] = str(j)
	return titles_mapping

def getIMDBPageLink(title):
	soup = getWebData(BASE_URL+'/find?q=' + title + '&s=tt')
	try:
		listing = soup.find('table').find_all('td',{'class':'result_text'})
	except:
		return None
	if len(listing) > 0:
		listing = listing[0]
	try:
		listing = listing.a['href']
	except:
		listing = None
	return listing

def getCompanyDetails(pageLink):
	company_details = {}
	company_details['production'] = []
	company_details['distributors'] = []
	soup = getWebData(BASE_URL+pageLink)
	both_list = soup.find_all('ul',{'class':'simpleList'})
	if len(both_list) > 0:
		company_details['production'] = [remWS(x.text) for x in both_list[0].find_all('li')]
	if len(both_list) > 1:
		company_details['distributors'] = [remWS(x.text) for x in both_list[1].find_all('li')]
	return company_details

def scrapeAndAdd(pageLink,sheet,rowNum,colMap):
	if pageLink is None:
		return
	soup = getWebData(BASE_URL+pageLink)
	titleBar = soup.find('div',{'class':'titleBar'})
	genres = ', '.join([x.text for x in titleBar.find_all('span',{'itemprop':'genre'})])
	runtime =  titleBar.find('time')
	try:
		category = titleBar.find('a',{'title':'See more release dates'}).text.strip()
		if 'series' in category.lower():
			category = 'TV Series'
		else:
			category = 'TV Movie'
	except:
		category = None
	release_date = ''
	country = ''
	language = ''
	if soup.find('div',id='titleDetails') is not None:
		for x in soup.find('div',id='titleDetails').find_all('h4'):
			if x.text == 'Release Date:':
				release_date = x.nextSibling.strip()
			if x.text == 'Country:':
				country = x.parent.text.replace('Country:','').replace('\n','')
			if x.text == 'Language:':
				language = x.parent.text.replace('Language:','').replace('\n','')
	company_details = getCompanyDetails(pageLink.replace('?','companycredits?'))
	producer = ', '.join(company_details['production'])
	distributors = ', '.join(company_details['distributors'])
	try:
		episodes = soup.find('a',{'class':'np_episode_guide'}).find('span').text.replace('episodes','').strip()
	except:
		episodes = None
	try:
		seasons = soup.find('div',{'class':'bp_text_only'}).find('div',{'class':'bp_heading'}).text
		seasons = re.findall('\d+', seasons)[0]
	except:
		seasons = None
	if runtime is not None:
		runtime = runtime.text.strip()
	ws[colMap['seasons']+rowNum].value = seasons
	ws[colMap['category']+rowNum].value = category
	ws[colMap['genre']+rowNum].value = genres
	ws[colMap['runtime']+rowNum].value = runtime
	ws[colMap['episodes']+rowNum].value = episodes
	ws[colMap['release']+rowNum].value = release_date
	ws[colMap['producer']+rowNum].value = producer
	ws[colMap['distributor']+rowNum].value = distributors
	ws[colMap['country']+rowNum].value = country
	ws[colMap['language']+rowNum].value = language


ws = wb[sheetName]
titlesMap = getTitles(ws)
columnMap = getColumnMap(ws)
for title in titlesMap.keys():
	print 'Scraping data for: ' + title
	scrapeAndAdd(getIMDBPageLink(title),ws,titlesMap[title],columnMap)
	wb.save(excelFile)

print("\n\nExecution Time: --- %2.f seconds ---\n" % (time.time() - start_time))
