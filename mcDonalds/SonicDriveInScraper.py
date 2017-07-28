from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
import requests, time, re, json, csv
import unicodedata

def getWebData(link):
    return BeautifulSoup(requests.get(link).text,'html.parser')

def remSpCh(s):
	if s is None:
		return ''
	s = unicodedata.normalize('NFKD', s).encode('ascii','ignore')
	return ''.join([i for i in s if ord(i) < 128])

def getStartIndex(sheet):
	for i in range(1,sheet.max_row+1):
		val = sheet['B'+str(i)].value
		if val is None:
			return i
	return i

def soupTextIfNotNone(obj):
	if obj is not None:
		return remSpCh(obj.text)
	return remSpCh(obj)

def getSheet(wb,sheetname):
	if sheetname in wb.get_sheet_names():
		return wb[sheetname]
	ws = wb.create_sheet()
	ws.title = sheetname
	return ws

def extractDetails(soup, city, state):
	restaurantName = soupTextIfNotNone(soup.find('h1',{'class':'nap-content-title'}))
	addressInfo = soup.find('div',{'class':'nap-content-address'})
	address = soupTextIfNotNone(addressInfo.find('span',{'class':'c-address-street-1'}))
	zipcode = soupTextIfNotNone(addressInfo.find('span',{'class':'c-address-postal-code'}))
	return [restaurantName, address, city, state, zipcode]

def updateExcelFile(wb, filename, data):
	ws = getSheet(wb, 'Dashboard')
	startIndex = getStartIndex(ws)
	if startIndex == 1:
		ws['A'+str(startIndex)].value = "Date Added"
		ws['B'+str(startIndex)].value = "Restaurant Name"
		ws['C'+str(startIndex)].value = "Street"
		ws['D'+str(startIndex)].value = "City"
		ws['E'+str(startIndex)].value = "State"
		ws['F'+str(startIndex)].value = "Zipcode"
		startIndex += 1
	for i, row in enumerate(data):
		ws['A'+str(startIndex+i)].value = date_today
		ws['B'+str(startIndex+i)].value = row[0]
		ws['C'+str(startIndex+i)].value = row[1]
		ws['D'+str(startIndex+i)].value = row[2]
		ws['E'+str(startIndex+i)].value = row[3]
		ws['F'+str(startIndex+i)].value = row[4]
	wb.save(filename)


start_time = time.time()
date_today = time.strftime("%m/%d/%y")
BASE_URL = "https://locations.sonicdrivein.com/"
excelFile = "Sonic DriveIn.xlsx"
try:
	wb = load_workbook(filename = excelFile,data_only=True)
except:
	wb = Workbook()
soup = getWebData(BASE_URL+'index.html')
for state in soup.find_all('li',{'class':'c-directory-list-content-item'}):
	soup = getWebData(BASE_URL + state.a['href'])
	print "Scraping " + state.a.text
	big_data = []
	for city in soup.find_all('li',{'class':'c-directory-list-content-item'}):
		soup = getWebData(BASE_URL + city.a['href'])
		count = city.find('span',{'class':'c-directory-list-content-item-count'}).text
		print "\tScraping " + city.a.text
		if count != '1':
			for location in soup.find_all('a',{'class':' c-location-grid-item-link-visit'}):
				soup = getWebData(BASE_URL + location['href'].replace('../',''))
				big_data.append(extractDetails(soup, city.a.text, state.a.text))
		else:
			big_data.append(extractDetails(soup, city.a.text, state.a.text))
	updateExcelFile(wb, excelFile, big_data)


print("\n\nExecution Time: --- %2.f seconds ---\n" % (time.time() - start_time))